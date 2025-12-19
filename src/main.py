import os
import json
import re
import logging
import time
import traceback
from pathlib import Path

import docx
import openai
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# Configuração de Logging
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')

# ======================
# CONFIGURAÇÃO GERAL
# ======================
# Use OPENAI_API_KEY by default. You can also set GITHUB_TOKEN for other providers.
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY") or os.getenv("GITHUB_TOKEN")
AI_PROVIDER = os.getenv("AI_PROVIDER", "openai").lower()  # 'openai' (default) or 'copilot' (custom)
DOCUMENTS_DIR = os.getenv("DOCUMENTS_DIR", "Documentações")
EXCEL_OUTPUT = os.getenv("EXCEL_OUTPUT", "cenarios_de_testes.xlsx")

MANUAL_MIN_PER_TEST = int(os.getenv("MANUAL_MIN_PER_TEST", "15"))
AI_MIN_PER_TEST = int(os.getenv("AI_MIN_PER_TEST", "2"))

# ======================
# CONFIGURAÇÃO OPENAI
# ======================
if AI_PROVIDER == "openai":
    if not OPENAI_API_KEY:
        logging.warning("OPENAI_API_KEY não encontrado nas variáveis de ambiente. A chamada à API falhará sem a chave.")
    else:
        openai.api_key = OPENAI_API_KEY

# Endpoint override (if needed)
AI_API_URL = os.getenv("AI_API_URL", "")

# ======================
# MOTOR DE IA
# ======================
def call_ai(prompt: str, max_retries: int = 3, backoff_base: float = 2.0) -> str:
    """Chama o provedor de IA configurado. Retorna conteúdo textual bruto da resposta."""
    if AI_PROVIDER == "openai":
        if not openai.api_key:
            raise RuntimeError("OPENAI_API_KEY não configurado nas variáveis de ambiente.")

        attempt = 0
        while True:
            try:
                logging.info("🔎 Chamando API do OpenAI...")
                # Use the Python SDK
                resp = openai.ChatCompletion.create(
                    model="gpt-4",
                    messages=[
                        {"role": "system", "content": "Você é um Engenheiro de QA Sênior focado em precisão técnica e JSON estruturado."},
                        {"role": "user", "content": prompt}
                    ],
                    temperature=0.1,
                    max_tokens=3000,
                )
                return resp['choices'][0]['message']['content']
            except Exception as e:
                attempt += 1
                logging.error(f"Erro na API (tentativa {attempt}): {e}")
                if attempt >= max_retries:
                    logging.debug(traceback.format_exc())
                    raise
                sleep_time = backoff_base ** attempt
                logging.info(f"Aguardando {sleep_time}s antes de nova tentativa...")
                time.sleep(sleep_time)
    else:
        # Placeholder for other providers (ex: GitHub Copilot) - custom implementation required
        raise NotImplementedError(f"AI_PROVIDER '{AI_PROVIDER}' não implementado. Use 'openai'.")

# =============================================
# ENGENHARIA DE PROMPT
# =============================================
QA_PROMPT_TEMPLATE = """
Aja como um Lead QA Engineer. Analise os requisitos abaixo e gere um Plano de Testes em formato JSON.

### REGRAS DE OUTPUT:
1. Retorne EXCLUSIVAMENTE o objeto JSON.
2. Não utilize blocos de código markdown (```json ... ```).
3. Use a técnica de 'Análise de Valor Limite' e 'Transição de Estados'.
4. IDs: TC-FUNC-NNN, TC-NEG-NNN, TC-SEC-NNN.

### ESTRUTURA DO JSON:
{
  "analise_requisitos": { "riscos": [], "entidades": [] },
  "cenarios_funcionais": [
    {
      "id": "TC-FUNC-001",
      "titulo": "Título Curto",
      "prioridade": "Alta",
      "descricao": "O que o teste faz",
      "passos": ["1...", "2..."],
      "resultado_esperado": "Resultado verificável"
    }
  ],
  "cenarios_negativos": [],
  "cenarios_borda": [],
  "metricas_qualidade": { "total_casos": 0 }
}

### REQUISITOS PARA ANÁLISE:
{requisitos_texto}
"""

# ======================
# PROCESSAMENTO DE ARQUIVOS
# ======================

def extrair_requisitos_docx(caminho: Path) -> str:
    logging.info(f"📄 Extraindo texto de: {caminho}")
    doc = docx.Document(str(caminho))
    conteudo = [para.text.strip() for para in doc.paragraphs if para.text.strip()]
    return "\n".join(conteudo)


def limpar_e_validar_json(texto: str, doc_stem: str = "response") -> dict:
    texto_limpo = re.sub(r'```json\s*|```', '', texto).strip()
    try:
        return json.loads(texto_limpo)
    except json.JSONDecodeError:
        # tenta extrair o primeiro objeto JSON
        match = re.search(r'\{.*\}', texto_limpo, re.DOTALL)
        if match:
            try:
                return json.loads(match.group(0))
            except json.JSONDecodeError:
                pass
        # salva a resposta bruta para inspeção
        logs_dir = Path(".logs")
        logs_dir.mkdir(exist_ok=True)
        raw_path = logs_dir / f"raw_response_{doc_stem}.txt"
        with raw_path.open("w", encoding="utf-8") as f:
            f.write(texto)
        raise RuntimeError(f"Falha ao converter resposta em JSON. Resposta bruta salva em: {raw_path}")


def salvar_excel_por_documentos(doc_json_list):
    logging.info("📝 Gerando arquivo Excel final...")
    wb = Workbook()
    # Remove aba padrão se estiver vazia
    if wb.active and wb.active.title == 'Sheet' and wb.active.max_row == 1 and wb.active.max_column == 1 and wb.active.cell(1, 1).value is None:
        wb.remove(wb.active)

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for doc_name, data in doc_json_list:
        ws = wb.create_sheet(title=doc_name[:30])
        headers = ["ID", "Título", "Prioridade", "Descrição", "Passos", "Resultado Esperado"]
        
        # Estilização do cabeçalho
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Unificar cenários
        todos_cenarios = data.get("cenarios_funcionais", []) + \
                         data.get("cenarios_negativos", []) + \
                         data.get("cenarios_borda", [])

        for row_num, cenario in enumerate(todos_cenarios, 2):
            ws.cell(row=row_num, column=1, value=cenario.get("id"))
            ws.cell(row=row_num, column=2, value=cenario.get("titulo"))
            ws.cell(row=row_num, column=3, value=cenario.get("prioridade"))
            ws.cell(row=row_num, column=4, value=cenario.get("descricao"))
            ws.cell(row=row_num, column=5, value="\n".join(cenario.get("passos", [])))
            ws.cell(row=row_num, column=6, value=cenario.get("resultado_esperado"))
            
            # Aplicar bordas e alinhamento
            for col in range(1, 7):
                ws.cell(row=row_num, column=col).border = thin_border
                ws.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True, vertical="top")

        # Ajuste de largura das colunas
        try:
            ws.column_dimensions['D'].width = 40
            ws.column_dimensions['E'].width = 50
            ws.column_dimensions['F'].width = 40
        except Exception:
            pass

    wb.save(EXCEL_OUTPUT)
    logging.info(f"✨ Sucesso! Planilha gerada: {EXCEL_OUTPUT}")

# ======================
# EXECUÇÃO PRINCIPAL
# ======================

def iniciar_testforge():
    # Garante que a pasta de documentações existe
    Path(DOCUMENTS_DIR).mkdir(exist_ok=True)
    
    docx_files = list(Path(DOCUMENTS_DIR).glob("*.docx"))
    
    if not docx_files:
        logging.warning(f"⚠️ Nenhum arquivo .docx encontrado na pasta '{DOCUMENTS_DIR}'. Coloque ao menos um arquivo .docx para gerar cenários.")
        return

    doc_json_list = []
    for docpath in docx_files:
        try:
            texto = extrair_requisitos_docx(docpath)
            prompt = QA_PROMPT_TEMPLATE.format(requisitos_texto=texto)
            resposta = call_ai(prompt)
            dados_qa = limpar_e_validar_json(resposta, docpath.stem)
            doc_json_list.append((docpath.stem, dados_qa))
        except Exception as e:
            logging.error(f"❌ Erro ao processar {docpath.name}: {e}")

    if doc_json_list:
        salvar_excel_por_documentos(doc_json_list)


if __name__ == "__main__":
    print("🚀 TestForge iniciado...")
    iniciar_testforge()
